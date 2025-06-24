"""
法律法规台账生成器
支持生成Excel、CSV、HTML格式的台账
"""
import os
from datetime import datetime
from typing import List, Optional, Dict, Any
import pandas as pd
from pathlib import Path
import logging
from jinja2 import Template
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

from ..storage.database import DatabaseManager
from ..storage.models import LawMetadata, LawStatus

logger = logging.getLogger(__name__)


class LedgerGenerator:
    """法律法规台账生成器"""
    
    def __init__(self, db_manager: Optional[DatabaseManager] = None):
        """
        初始化台账生成器
        
        Args:
            db_manager: 数据库管理器实例
        """
        self.db_manager = db_manager or DatabaseManager()
        self.output_dir = Path("data/ledgers")
        self.output_dir.mkdir(parents=True, exist_ok=True)
        
    def generate_ledger(self, 
                       output_format: str = "excel",
                       filename: Optional[str] = None,
                       filters: Optional[Dict[str, Any]] = None) -> str:
        """
        生成法律法规台账
        
        Args:
            output_format: 输出格式 (excel, csv, html)
            filename: 输出文件名（不含扩展名）
            filters: 过滤条件
            
        Returns:
            生成的文件路径
        """
        # 获取数据
        laws = self._fetch_law_data(filters)
        
        if not laws:
            logger.warning("没有找到任何法律法规数据")
            return None
            
        # 准备台账数据
        ledger_data = self._prepare_ledger_data(laws)
        
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            filename = f"law_ledger_{timestamp}"
            
        # 根据格式生成台账
        if output_format.lower() == "excel":
            return self._generate_excel(ledger_data, filename)
        elif output_format.lower() == "csv":
            return self._generate_csv(ledger_data, filename)
        elif output_format.lower() == "html":
            return self._generate_html(ledger_data, filename)
        else:
            raise ValueError(f"不支持的输出格式: {output_format}")
            
    def _fetch_law_data(self, filters: Optional[Dict[str, Any]] = None) -> List[LawMetadata]:
        """
        从数据库获取法律法规数据
        
        Args:
            filters: 过滤条件
            
        Returns:
            法律法规列表
        """
        with self.db_manager.get_session() as session:
            query = session.query(LawMetadata)
            
            if filters:
                # 应用过滤条件
                if 'status' in filters:
                    query = query.filter(LawMetadata.status == filters['status'])
                if 'law_type' in filters:
                    query = query.filter(LawMetadata.law_type == filters['law_type'])
                if 'valid_from' in filters:
                    query = query.filter(LawMetadata.valid_from >= filters['valid_from'])
                if 'valid_to' in filters:
                    query = query.filter(LawMetadata.valid_to <= filters['valid_to'])
                    
            # 按生效日期排序
            laws = query.order_by(LawMetadata.valid_from.desc()).all()
            
            # 确保会话中的对象在返回前被完全加载
            return [self._detach_law(law) for law in laws]
            
    def _detach_law(self, law: LawMetadata) -> Dict[str, Any]:
        """将SQLAlchemy对象转换为字典"""
        return {
            'name': law.name,
            'number': law.number,
            'law_type': law.law_type,
            'issuing_authority': law.issuing_authority,
            'publish_date': law.publish_date,
            'valid_from': law.valid_from,
            'valid_to': law.valid_to,
            'status': law.status,
            'source': law.source,
            'source_url': law.source_url,
            'created_at': law.created_at,
            'updated_at': law.updated_at
        }
            
    def _prepare_ledger_data(self, laws: List[Dict[str, Any]]) -> pd.DataFrame:
        """
        准备台账数据
        
        Args:
            laws: 法律法规列表
            
        Returns:
            DataFrame格式的台账数据
        """
        ledger_records = []
        
        for law in laws:
            record = {
                '序号': len(ledger_records) + 1,
                '法规名称': law['name'],
                '文号': law['number'] or '',
                '发布机关': law['issuing_authority'] or '',
                '发布日期': law['publish_date'].strftime('%Y-%m-%d') if law['publish_date'] else '',
                '生效日期': law['valid_from'].strftime('%Y-%m-%d') if law['valid_from'] else '',
                '失效日期': law['valid_to'].strftime('%Y-%m-%d') if law['valid_to'] else '',
                '状态': self._get_status_text(law['status']),
                '数据来源': law['source'] or '',
                '来源链接': law['source_url'] or '',  # 添加来源链接字段
                '采集时间': law['created_at'].strftime('%Y-%m-%d %H:%M:%S') if law['created_at'] else '',
                '更新时间': law['updated_at'].strftime('%Y-%m-%d %H:%M:%S') if law['updated_at'] else ''
            }
            ledger_records.append(record)
            
        return pd.DataFrame(ledger_records)
        
    def _get_status_text(self, status: str) -> str:
        """获取状态的中文描述"""
        status_map = {
            LawStatus.DRAFT: "草案",
            LawStatus.EFFECTIVE: "有效",
            LawStatus.AMENDED: "已修正",
            LawStatus.REPEALED: "已废止",
            LawStatus.EXPIRED: "已失效"
        }
        # 统一状态字段 - 将"现行有效"统一为"有效"
        if status in ["现行有效", "有效"]:
            return "有效"
        return status_map.get(status, status)
        
    def _generate_excel(self, data: pd.DataFrame, filename: str) -> str:
        """
        生成Excel格式的台账
        
        Args:
            data: 台账数据
            filename: 文件名
            
        Returns:
            文件路径
        """
        filepath = self.output_dir / f"{filename}.xlsx"
        
        # 创建Excel writer
        with pd.ExcelWriter(filepath, engine='openpyxl') as writer:
            # 写入数据
            data.to_excel(writer, sheet_name='法律法规台账', index=False)
            
            # 获取工作表
            workbook = writer.book
            worksheet = writer.sheets['法律法规台账']
            
            # 设置样式
            self._style_excel_sheet(worksheet, len(data))
            
        logger.info(f"Excel台账已生成: {filepath}")
        return str(filepath)
        
    def _style_excel_sheet(self, worksheet, row_count: int):
        """
        设置Excel工作表样式
        
        Args:
            worksheet: 工作表对象
            row_count: 数据行数
        """
        # 标题行样式
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_alignment = Alignment(horizontal="center", vertical="center")
        
        # 边框样式
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # 应用标题行样式
        for cell in worksheet[1]:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
            
        # 设置列宽 - 更新为12列（增加了来源链接）
        column_widths = {
            'A': 8,   # 序号
            'B': 50,  # 法规名称
            'C': 30,  # 文号
            'D': 20,  # 发布机关
            'E': 15,  # 发布日期
            'F': 15,  # 生效日期
            'G': 15,  # 失效日期
            'H': 12,  # 状态
            'I': 20,  # 数据来源
            'J': 40,  # 来源链接
            'K': 20,  # 采集时间
            'L': 20   # 更新时间
        }
        
        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width
            
        # 应用数据行样式和超链接
        for row in range(2, row_count + 2):
            for col in range(1, 13):  # 更新为13列
                cell = worksheet.cell(row=row, column=col)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
                
                # 处理来源链接列（第J列）的超链接
                if col == 10:  # 来源链接列
                    cell_value = cell.value
                    if cell_value and cell_value.startswith('http'):
                        cell.hyperlink = cell_value
                        cell.style = "Hyperlink"
                
        # 冻结首行
        worksheet.freeze_panes = 'A2'
        
    def _generate_csv(self, data: pd.DataFrame, filename: str) -> str:
        """
        生成CSV格式的台账
        
        Args:
            data: 台账数据
            filename: 文件名
            
        Returns:
            文件路径
        """
        filepath = self.output_dir / f"{filename}.csv"
        data.to_csv(filepath, index=False, encoding='utf-8-sig')
        logger.info(f"CSV台账已生成: {filepath}")
        return str(filepath)
        
    def _generate_html(self, data: pd.DataFrame, filename: str) -> str:
        """
        生成HTML格式的台账
        
        Args:
            data: 台账数据
            filename: 文件名
            
        Returns:
            文件路径
        """
        filepath = self.output_dir / f"{filename}.html"
        
        # HTML模板
        html_template = """
<!DOCTYPE html>
<html lang="zh-CN">
<head>
    <meta charset="UTF-8">
    <meta name="viewport" content="width=device-width, initial-scale=1.0">
    <title>法律法规台账</title>
    <style>
        body {
            font-family: -apple-system, BlinkMacSystemFont, "Segoe UI", Roboto, "Helvetica Neue", Arial, sans-serif;
            padding: 20px;
            background-color: #f5f5f5;
        }
        .container {
            max-width: 100%;
            margin: 0 auto;
            background-color: white;
            padding: 20px;
            border-radius: 8px;
            box-shadow: 0 2px 4px rgba(0,0,0,0.1);
        }
        h1 {
            color: #333;
            text-align: center;
            margin-bottom: 30px;
        }
        .info {
            margin-bottom: 20px;
            padding: 10px;
            background-color: #f0f8ff;
            border-radius: 4px;
        }
        table {
            width: 100%;
            border-collapse: collapse;
            margin-top: 20px;
        }
        th {
            background-color: #366092;
            color: white;
            padding: 12px;
            text-align: left;
            font-weight: bold;
            position: sticky;
            top: 0;
            z-index: 10;
        }
        td {
            padding: 10px;
            border-bottom: 1px solid #ddd;
        }
        tr:hover {
            background-color: #f5f5f5;
        }
        .status-effective {
            color: #28a745;
            font-weight: bold;
        }
        .status-repealed {
            color: #dc3545;
            font-weight: bold;
        }
        .status-amended {
            color: #ffc107;
            font-weight: bold;
        }
        .status-expired {
            color: #6c757d;
            font-weight: bold;
        }
        @media print {
            .no-print {
                display: none;
            }
        }
    </style>
</head>
<body>
    <div class="container">
        <h1>法律法规台账</h1>
        <div class="info">
            <p><strong>生成时间：</strong>{{ generation_time }}</p>
            <p><strong>记录总数：</strong>{{ total_count }} 条</p>
        </div>
        <table>
            <thead>
                <tr>
                    {% for col in columns %}
                    <th>{{ col }}</th>
                    {% endfor %}
                </tr>
            </thead>
            <tbody>
                {% for row in rows %}
                <tr>
                    {% for col in columns %}
                    <td{% if col == '状态' %} class="status-{{ row[col]|lower|replace(' ', '-') }}"{% endif %}>
                        {{ row[col] }}
                    </td>
                    {% endfor %}
                </tr>
                {% endfor %}
            </tbody>
        </table>
    </div>
</body>
</html>
        """
        
        # 准备模板数据
        template = Template(html_template)
        html_content = template.render(
            generation_time=datetime.now().strftime('%Y-%m-%d %H:%M:%S'),
            total_count=len(data),
            columns=data.columns.tolist(),
            rows=data.to_dict('records')
        )
        
        # 写入文件
        with open(filepath, 'w', encoding='utf-8') as f:
            f.write(html_content)
            
        logger.info(f"HTML台账已生成: {filepath}")
        return str(filepath)
        
    def generate_summary_report(self) -> Dict[str, Any]:
        """
        生成汇总报告
        
        Returns:
            汇总统计信息
        """
        with self.db_manager.get_session() as session:
            total_count = session.query(LawMetadata).count()
            
            # 按状态统计
            status_stats = {}
            for status in LawStatus:
                count = session.query(LawMetadata).filter(
                    LawMetadata.status == status.value
                ).count()
                status_stats[self._get_status_text(status.value)] = count
                
            # 按类型统计
            type_stats = {}
            law_types = session.query(LawMetadata.law_type).distinct().all()
            for law_type, in law_types:
                if law_type:
                    count = session.query(LawMetadata).filter(
                        LawMetadata.law_type == law_type
                    ).count()
                    type_stats[law_type] = count
                    
            # 按年份统计
            year_stats = {}
            laws_with_date = session.query(LawMetadata).filter(
                LawMetadata.valid_from.isnot(None)
            ).all()
            
            for law in laws_with_date:
                year = law.valid_from.year
                year_stats[year] = year_stats.get(year, 0) + 1
                
        return {
            'total_count': total_count,
            'status_stats': status_stats,
            'type_stats': type_stats,
            'year_stats': dict(sorted(year_stats.items(), reverse=True)[:10])  # 最近10年
        } 